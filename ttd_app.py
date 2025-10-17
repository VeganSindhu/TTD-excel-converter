import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

# Title
st.title("TTD Excel Processor")

# File uploader
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    file_bytes = uploaded_file.getvalue()
    
    try:
        # Load workbook for accurate row reading
        wb_buffer = io.BytesIO(file_bytes)
        wb = load_workbook(wb_buffer, data_only=True)
        ws = wb.active

        # Read first 3 rows using openpyxl for precision
        data_headers = [cell.value for cell in ws[1]]
        mapping = [cell.value for cell in ws[2]]
        desired_headers = [cell.value for cell in ws[3]]

        # Pad to max length
        max_len = max(len(data_headers), len(mapping), len(desired_headers))
        data_headers += [None] * (max_len - len(data_headers))
        mapping += [None] * (max_len - len(mapping))
        desired_headers += [None] * (max_len - len(desired_headers))

        # Read data from row 4 onwards using pandas
        data_buffer = io.BytesIO(file_bytes)
        data_df = pd.read_excel(data_buffer, header=None, skiprows=3)
        data_df.columns = data_headers[:data_df.shape[1]]

        # Updated function to split address
        def split_address(addr):
            if pd.isna(addr):
                return ['', '', '', '', '', '']
            s = str(addr)
            parts = [p.strip() for p in s.split(',') if p.strip()]
            if len(parts) >= 3:
                pin = parts[-1]
                state = parts[-2]
                city = parts[-3]
                addr_parts = parts[:-3]
            elif len(parts) == 2:
                pin = parts[-1]
                state = parts[0]
                city = ''
                addr_parts = []
            elif len(parts) == 1:
                pin = parts[0]
                state = ''
                city = ''
                addr_parts = []
            else:
                pin = ''
                state = ''
                city = ''
                addr_parts = []
            
            # Assign lines with logic for gaps and remaining parts
            line1 = addr_parts[0] if len(addr_parts) > 0 else ''
            line2 = addr_parts[1] if len(addr_parts) > 1 else city  # Fill with city if no line2
            line3 = ', '.join(addr_parts[2:]) if len(addr_parts) > 2 else city  # Join 3rd + all remaining; fill with city if none
            return [line1, line2, line3, city, state, pin]

        # Apply address split if 'Address' column exists
        if 'Address' in data_df.columns:
            split_results = data_df['Address'].apply(split_address)
            split_df = pd.DataFrame(split_results.tolist(), columns=['line1', 'line2', 'line3', 'city_from_addr', 'state_from_addr', 'pin_from_addr'], index=data_df.index)
            data_df = pd.concat([data_df, split_df], axis=1)
        else:
            data_df[['line1', 'line2', 'line3', 'city_from_addr', 'state_from_addr', 'pin_from_addr']] = ['', '', '', '', '', '']

        # Build output DataFrame
        output_df = pd.DataFrame()

        clean_data_headers = {str(h).strip().lower(): h for h in data_headers if h is not None}

        for i, desired in enumerate(desired_headers):
            if desired is None:
                continue

            map_item = mapping[i]

            # Special handling (override mapping for these)
            desired_lower = str(desired).strip().lower()
            if 'receiver add line 1' in desired_lower:
                output_df[desired] = data_df['line1']
                continue
            elif 'receiver add line 2' in desired_lower:
                output_df[desired] = data_df['line2']
                continue
            elif 'receiver add line 3' in desired_lower:
                output_df[desired] = data_df['line3']
                continue
            elif 'receiver state/ut' in desired_lower:
                output_df[desired] = data_df['state_from_addr']
                continue
            elif 'receiver pincode' in desired_lower:
                output_df[desired] = data_df['PinCode'] if 'PinCode' in data_df.columns else data_df['pin_from_addr']
                continue
            elif 'receiver city' in desired_lower:
                output_df[desired] = data_df['City'] if 'City' in data_df.columns else data_df['city_from_addr']
                continue
            elif 'sender add line 1' in desired_lower:
                output_df[desired] = ['SALES WING OF PUBLICATIONS'] * len(data_df)
                continue
            elif 'sender add line 2' in desired_lower:
                output_df[desired] = ['TTD PRESS COMPOUND'] * len(data_df)
                continue
            elif 'sender add line 3' in desired_lower:
                output_df[desired] = ['Tirupati - 517507'] * len(data_df)
                continue

            # Normal mapping or constant
            if map_item is not None:
                map_str = str(map_item).strip().lower()
                if map_str in clean_data_headers:
                    orig_col = clean_data_headers[map_str]
                    output_df[desired] = data_df[orig_col]
                else:
                    # Constant: repeat map_item for all rows
                    output_df[desired] = [map_item] * len(data_df)
            else:
                output_df[desired] = [''] * len(data_df)

        # Reorder columns to match desired_headers order
        output_cols = [col for col in desired_headers if col in output_df.columns]
        output_df = output_df[output_cols]

        # Write to in-memory Excel
        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            # First row: desired headers
            pd.DataFrame([desired_headers]).to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=0)
            # Data from second row
            output_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=1)

        output_buffer.seek(0)

        st.download_button(
            label="Download Processed Excel",
            data=output_buffer.getvalue(),
            file_name="ttd_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success("✅ File processed successfully!")
        
    except Exception as e:
        st.error(f"❌ Processing error: {str(e)}")
        import traceback
        st.code(traceback.format_exc())